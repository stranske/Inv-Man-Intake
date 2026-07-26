"""Return-series spreadsheet export coverage."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from math import inf, nan

import pytest
from openpyxl import load_workbook

from inv_man_intake.export import export_return_series
from inv_man_intake.extraction.providers import ExtractedTable, ExtractedTableCell, SourceLocation
from inv_man_intake.performance import PerformancePayload, PerformancePoint, PerformanceSeries


def test_return_series_roundtrips_with_provenance() -> None:
    payload = PerformancePayload(
        monthly=PerformanceSeries(
            frequency="monthly",
            points=(
                PerformancePoint(as_of=date(2026, 1, 31), value=0.125),
                PerformancePoint(as_of=date(2026, 2, 28), value=-0.01),
            ),
        )
    )
    manifest = export_return_series(
        payload,
        source_doc_id="track-record.pdf",
        source_page=4,
        method="performance-normalizer",
        tables=(
            ExtractedTable(
                table_id="returns",
                location=SourceLocation(source_doc_id="track-record.pdf", source_page=4),
                cells=(ExtractedTableCell(row_index=0, column_index=0, value="Jan 2026"),),
            ),
        ),
    )

    artifacts = {entry.item_ref: entry.artifact for entry in manifest.artifacts}
    workbook = load_workbook(BytesIO(artifacts["return-series.xlsx"].content), data_only=True)
    assert workbook.sheetnames == ["Return Series", "Manifest"]
    assert list(workbook["Return Series"].iter_rows(values_only=True)) == [
        ("period", "value", "frequency", "provenance"),
        ("2026-01-31", 0.125, "monthly", "track-record.pdf:4:performance-normalizer"),
        ("2026-02-28", -0.01, "monthly", "track-record.pdf:4:performance-normalizer"),
    ]
    assert any(
        row[0] == "table_cell" and "Jan 2026" in row[1]
        for row in workbook["Manifest"].iter_rows(min_row=2, values_only=True)
    )
    assert b"track-record.pdf:4:performance-normalizer" in artifacts["return-series.csv"].content


@pytest.mark.parametrize("bad_value", (nan, inf, -inf))
def test_non_finite_value_is_rejected_not_written(bad_value: float) -> None:
    payload = PerformancePayload(
        monthly=PerformanceSeries(
            frequency="monthly",
            points=(PerformancePoint(as_of=date(2026, 1, 31), value=bad_value),),
        )
    )
    with pytest.raises(ValueError, match="finite"):
        export_return_series(payload, source_doc_id="track-record.pdf")


def test_missing_series_records_skip() -> None:
    manifest = export_return_series(None, source_doc_id="track-record.pdf")
    assert [(entry.item_ref, entry.reason_code) for entry in manifest.skipped] == [
        ("return-series", "no_series_found")
    ]
